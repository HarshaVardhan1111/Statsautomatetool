from flask import Flask, request, send_file, render_template
import openpyxl
import re
import os
app = Flask(__name__)
# Function to clean Emp ID by removing all non-printable characters and extra spaces
def clean_emp_id(emp_id):
   if emp_id is None:
       return None
   emp_id = str(emp_id).strip()
   emp_id = re.sub(r'\s+', '', emp_id)
   emp_id = ''.join(c for c in emp_id if c.isprintable())
   return emp_id.upper()
@app.route('/')
def index():
   return render_template('index.html')
@app.route('/upload', methods=['POST'])
def upload_file():
   if 'file' not in request.files:
       return 'No file part', 400
   file = request.files['file']
   if file.filename == '':
       return 'No selected file', 400
   if file and file.filename.endswith('.xlsx'):
       jul_file_path = os.path.join('uploads', file.filename)
       file.save(jul_file_path)
       final_file_path = r'C:\Users\hsubra154\Downloads\Final File (1).xlsx'  # Set your final file path
       # Load the files
       jul_wb = openpyxl.load_workbook(jul_file_path)
       final_wb = openpyxl.load_workbook(final_file_path)
       # Define your sheet mappings here
       sheet_row_mapping = {
    'Ad Copy': (3, 15),
    'Retail Ad Copy': (16, 28),
    'LA Uploads': (29, 37),
    'Retail Uploads': (38, 46),
    'Coding and Uploads': (47, 54),
    'Enterprise QC': (55, 67),
    'Enterprise Uploads': (68, 76),
    'Retail CSM Coding': (77, 85),
    'DR': (86, 94),
    'Amp DR': (95, 103),
    'Amp OE': (104, 112),
    'ROE': (113, 121),
    'QAR': (122, 130),
    'MG': (131, 139),
}
       # Process each sheet in the Jul file
       for jul_sheet_name, (start_row, end_row) in sheet_row_mapping.items():
           if jul_sheet_name not in jul_wb.sheetnames:
               continue
           jul_ws = jul_wb[jul_sheet_name]
           emp_id_col = next((cell.column for cell in jul_ws[1] if cell.value == "Emp ID"), None)
           if emp_id_col is None:
               continue
           jul_data = {}
           for jul_row in jul_ws.iter_rows(min_row=2, min_col=1, max_col=14):
               cleaned_emp_id = clean_emp_id(jul_row[emp_id_col - 1].value)
               if cleaned_emp_id:
                   jul_data[cleaned_emp_id] = jul_row
           for final_sheet in final_wb.sheetnames:
               final_ws = final_wb[final_sheet]
               final_emp_id = clean_emp_id(final_ws['A2'].value)
               if final_emp_id and final_emp_id in jul_data:
                   jul_row_data = jul_data[final_emp_id]
                   for row_offset, jul_cell in enumerate(jul_row_data[2:15], start=start_row):
                       target_cell = final_ws.cell(row=row_offset, column=10)
                       target_cell.value = jul_cell.value
                       # Copy number formats
                       if isinstance(jul_cell.value, float):
                           if '%' in jul_cell.number_format:
                               target_cell.number_format = jul_cell.number_format
                           elif jul_cell.value.is_integer():
                               target_cell.number_format = '0'
                           else:
                               target_cell.number_format = '0.00'
                       else:
                           target_cell.number_format = 'General'
       output_file_path = 'output/sept1216.xlsx'
       final_wb.save(output_file_path)
       return send_file(output_file_path, as_attachment=True)
   return 'Invalid file format. Please upload an .xlsx file.', 400
if __name__ == '__main__':
   os.makedirs('uploads', exist_ok=True)
   os.makedirs('output', exist_ok=True)
   app.run(debug=True)